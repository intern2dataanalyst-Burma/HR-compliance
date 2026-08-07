import calendar
import io
import os
import re
import shutil
import smtplib
import tempfile
import zipfile
from copy import copy
from datetime import datetime
from email.message import EmailMessage
from io import BytesIO
from pathlib import Path

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter


custom_css = """
<style>
/* Global Background & Font */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');
html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
    background: radial-gradient(circle at top, #ffffff 0, #f3f6fb 45%, #e5edf7 100%) !important;
    color: #1e293b;
}

.block-container {
    padding-top: 1.5rem !important;
    padding-bottom: 2rem !important;
}

[data-testid="stDataFrame"] {
    background: #ffffff;
    border-radius: 12px;
    border: 1px solid #e5e7eb;
    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
}

.stDownloadButton > button {
    background: linear-gradient(to right, #1d9bf0, #34d399) !important;
    color: white !important;
    border: none !important;
    border-radius: 999px !important;
    font-weight: bold !important;
    padding: 0.5rem 1.5rem !important;
}
</style>
"""


BASE_DIR = Path(__file__).resolve().parent
DATA_CANDIDATES = [
    BASE_DIR / "data" / "real_master_compliance.xlsx",
    BASE_DIR / "sample_data" / "dummy_master_compliance.xlsx",
]
TEMPLATES_DIR = BASE_DIR / "templates"

# 1. Automatically create .streamlit/config.toml and enable static file serving if missing
config_dir = BASE_DIR / ".streamlit"
config_dir.mkdir(parents=True, exist_ok=True)
config_file = config_dir / "config.toml"
if not config_file.exists() or "enableStaticServing" not in config_file.read_text(encoding="utf-8"):
    with open(config_file, "a", encoding="utf-8") as f:
        f.write("\n[server]\nenableStaticServing = true\n")

# 2. Primary Archive for Web Linking (Streamlit Static HTTP Server)
STATIC_ARCHIVE_DIR = BASE_DIR / "static" / "HR_Compliance_Archive"
STATIC_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
STATIC_WEB_PREFIX = "app/static/HR_Compliance_Archive"

# 3. Secondary Archive for Windows OneDrive Cloud Sync
user_home = Path.home()
onedrive_path_secret = st.secrets.get("archive", {}).get("onedrive_path", None)
if onedrive_path_secret and Path(onedrive_path_secret).parent.exists():
    ONEDRIVE_ARCHIVE_DIR = Path(onedrive_path_secret)
else:
    onedrive_candidates = [
        user_home / "OneDrive - Hunger Pangs Pvt Ltd" / "HR_Compliance_Archive",
        user_home / "OneDrive" / "HR_Compliance_Archive",
    ]
    ONEDRIVE_ARCHIVE_DIR = next((path for path in onedrive_candidates if path.parent.exists()), None)

if ONEDRIVE_ARCHIVE_DIR:
    ONEDRIVE_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)


def email_file_to_outlook(file_bytes, filename):
    """
    Emails the generated file to Outlook with a structured subject tag 
    so Power Automate can automatically ingest it into OneDrive.
    """
    try:
        if hasattr(file_bytes, "getvalue"):
            file_bytes = file_bytes.getvalue()

        email_secrets = st.secrets.get("email_config", {})
        sender_email = email_secrets.get("sender_email")
        sender_password = email_secrets.get("sender_password")
        recipient_email = email_secrets.get("recipient_email", sender_email)

        if not sender_email or not sender_password:
            return False, "Email sender credentials missing in secrets.toml."

        msg = EmailMessage()
        msg["Subject"] = f"[HR Compliance Sync] {filename}"
        msg["From"] = sender_email
        msg["To"] = recipient_email
        msg.set_content(
            "Automated HR Compliance form generated from Streamlit portal. "
            "Attached is the file for Power Automate OneDrive archiving."
        )

        if filename.lower().endswith(".zip"):
            maintype, subtype = "application", "zip"
        else:
            maintype, subtype = (
                "application",
                "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        msg.add_attachment(
            file_bytes,
            maintype=maintype,
            subtype=subtype,
            filename=filename,
        )

        with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(sender_email, sender_password)
            smtp.send_message(msg)

        return True, "Emailed to Outlook for automatic OneDrive ingestion!"
    except Exception as e:
        return False, f"Email dispatch skipped/failed: {str(e)}"


@st.cache_data(show_spinner=False, ttl=300)
def normalize_unit_key(value: object) -> str:
    """Normalize a unit name for matching between Conso_Data and
    Units_Master. Verified against the real master file: the two sheets
    spell many unit names differently ('Mumbai Fort' vs 'Mumbai-Fort',
    'Gurugram Worldmark 65' vs 'Gurugram Worldmark65') - stripping all
    hyphens/underscores/whitespace, not just case, resolves those without
    needing the spreadsheet's unit names to be rewritten."""
    return re.sub(r"[\s\-_]+", "", str(value)).strip().lower()


def load_master_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    try:
        excel_url = st.secrets["EXCEL_FILE_URL"]
        response = requests.get(excel_url, timeout=15)
        response.raise_for_status()
        file_bytes = io.BytesIO(response.content)

        # Load data sheets (skipping row 1 because headers start on row 2)
        df_conso = pd.read_excel(file_bytes, sheet_name="Conso_Data", skiprows=1, engine="openpyxl")
        df_units = pd.read_excel(file_bytes, sheet_name="Units_Master", skiprows=1, engine="openpyxl")

        # Load mapping rule sheets directly from the same spreadsheet
        df_mapping_rules = pd.read_excel(file_bytes, sheet_name="Sheet1", header=None, engine="openpyxl")
        df_col_ref = pd.read_excel(file_bytes, sheet_name="Coloums", skiprows=3, engine="openpyxl")

        # The templates reference three more source registers ("Emp Mast",
        # "Leave Reg", "Attd Reg") that don't exist in the master workbook
        # yet. Load them if/when they're added, without crashing if absent -
        # Form C's leave columns and Form V's attendance grid will start
        # filling automatically as soon as these sheets appear.
        def _try_load_optional_sheet(candidate_names: list[str]) -> pd.DataFrame:
            for candidate in candidate_names:
                try:
                    file_bytes.seek(0)
                    sheet_df = pd.read_excel(file_bytes, sheet_name=candidate, engine="openpyxl")
                    sheet_df.columns = sheet_df.columns.astype(str).str.strip()
                    return sheet_df
                except (ValueError, KeyError):
                    continue
            return pd.DataFrame()

        df_emp_master = _try_load_optional_sheet(["Emp_Master", "Employee_Master", "EmpMaster", "Emp Mast"])
        df_leave_register = _try_load_optional_sheet(["Leave_Register", "Leave_Reg", "LeaveReg", "Leave Reg"])
        df_attendance_register = _try_load_optional_sheet(["Attendance_Register", "Attd_Reg", "AttdReg", "Attd Reg"])
        file_bytes.seek(0)

        df_conso.columns = df_conso.columns.astype(str).str.strip()
        df_units.columns = df_units.columns.astype(str).str.strip()

        if "Unit" in df_conso.columns:
            df_conso["Unit"] = df_conso["Unit"].astype(str).str.strip()
            df_conso["Unit_Display"] = df_conso["Unit"].astype(str).str.strip()
            df_conso["Unit_Clean"] = df_conso["Unit"].apply(normalize_unit_key)
        else:
            df_conso["Unit"] = pd.Series(["" for _ in range(len(df_conso))])
            df_conso["Unit_Display"] = pd.Series(["" for _ in range(len(df_conso))])
            df_conso["Unit_Clean"] = ""

        if "Unit" in df_units.columns:
            df_units["Unit"] = df_units["Unit"].astype(str).str.strip()
            df_units["Unit_Clean"] = df_units["Unit"].apply(normalize_unit_key)
        else:
            df_units["Unit"] = pd.Series(["" for _ in range(len(df_units))])
            df_units["Unit_Clean"] = ""

        if "Unit_Clean" in df_conso.columns and "Unit_Clean" in df_units.columns:
            state_lookup = df_units[["Unit_Clean", "State"]].drop_duplicates(subset=["Unit_Clean"])
            df_conso = df_conso.merge(state_lookup, on="Unit_Clean", how="left", suffixes=("", "_master"))
            if "State_master" in df_conso.columns:
                df_conso["State"] = df_conso["State"].fillna(df_conso["State_master"])
                df_conso.drop(columns=["State_master"], inplace=True)

        return df_conso, df_units, df_mapping_rules, df_col_ref, df_emp_master, df_leave_register, df_attendance_register
    except Exception as exc:
        st.error(f"Failed to download master data from Google Drive: {exc}")
        st.stop()


@st.cache_data(show_spinner=False)
def build_merged_view() -> pd.DataFrame:
    df_conso, df_units, df_mapping_rules, df_col_ref, df_emp_master, df_leave_register, df_attendance_register = load_master_data()
    df = df_conso.copy()

    # Fill in Father/Spouse Name, Sex, DOB etc. from the Employee Master when
    # Conso_Data doesn't carry them and Emp Mast has since been added.
    if not df_emp_master.empty and "Unit" in df.columns:
        code_col = next((c for c in df_emp_master.columns if "code" in c.lower()), None)
        conso_code_col = next((c for c in df.columns if "code" in c.lower()), None)
        if code_col and conso_code_col:
            df = df.merge(
                df_emp_master.add_suffix("_empmast").rename(columns={f"{code_col}_empmast": code_col}),
                left_on=conso_code_col,
                right_on=code_col,
                how="left",
                suffixes=("", "_empmast_dup"),
            )

    canonical_columns = {
        "Code": ["Empl_Code", "Code", "Employee Code", "Emp Code"],
        "Employee Name": ["Name_of_the_employee", "Employee Name", "Employee_Name", "Name"],
        "Father Name": ["Father_Name", "Father Name"],
        "Spouse Name": ["Spouce_Name", "Spouse Name", "Spouse_Name"],
        "UAN No": ["UAN No", "UAN", "UAN Number"],
        "PF No": ["PF No", "PF", "Provident Fund No"],
        "ESIC Old No": ["ESIC Old No", "ESIC No", "ESIC"],
        "PAN": ["PAN", "Pan"],
        "Joining_Date": ["Joining_Date", "Joining Date", "Join Date", "Joining_date"],
        "Exit_Date": ["Exit_Date", "Exit Date", "Exit_Date"],
        "Designation": ["Designation", "Designation Name"],
        "Department": ["Department", "Department Name"],
        "Days Paid": ["Days Paid", "Days_Paid"],
        "Days Present": ["Days Present", "Days_Present"],
        "Earned Basic": ["Earned Basic", "Earned_Basic", "Basic"],
        "Earned HRA": ["Earned HRA", "Earned_HRA", "HRA"],
        "Earned Gross Salary": ["Earned Gross Salary", "Earned_Gross_Salary", "Gross Salary", "Gross"],
        "Prov Fund": ["Prov Fund", "Prov_Fund", "PFund"],
        "ESIC": ["ESIC", "ESIC Amount"],
        "PTax": ["PTax", "Professional Tax"],
        "TDS": ["TDS", "Tax Deducted"],
        "Total Deductions": ["Total Deductions", "Total_Deductions", "Deduction", "Total Deduction"],
        "Net Paid": ["Net Paid", "Net_Paid", "Net Salary"],
    }

    # NOTE: Sheet1 is NOT a generic [target_name, alias1, alias2...] alias
    # table - it's the form-generation rule sheet (verified against the real
    # master file), with each row being either a form's header labels or its
    # rule values. Treating its rows as dashboard column aliases previously
    # corrupted "Unit" (and could corrupt any other real column whose name
    # happened to match a word in a form's header row, e.g. "Code").

    for canonical_name, aliases in canonical_columns.items():
        matched_column = next((alias for alias in aliases if alias in df.columns), None)
        if matched_column is not None:
            df[canonical_name] = df[matched_column]
        else:
            df[canonical_name] = pd.Series(["" for _ in range(len(df))])

    state_candidate = next((col for col in ["State", "Region", "Location", "State/Region", "State_Region"] if col in df.columns), None)
    if state_candidate is not None:
        df["State"] = df[state_candidate].astype(str).str.strip()
    else:
        df["State"] = pd.Series(["" for _ in range(len(df))])

    if "Unit" in df.columns:
        df["Unit_Raw"] = df["Unit"].astype(str).str.strip()
        df["Unit_Display"] = df["Unit_Raw"]
        df["Unit_Clean"] = df["Unit_Raw"].apply(normalize_unit_key)
    else:
        df["Unit_Raw"] = pd.Series(["" for _ in range(len(df))])
        df["Unit_Display"] = pd.Series(["" for _ in range(len(df))])
        df["Unit_Clean"] = pd.Series(["" for _ in range(len(df))])

    month_candidate = next((col for col in ["Month Year", "Month", "Pay Month", "Month_Year", "Period", "For the Month", "Date"] if col in df.columns), None)
    if month_candidate is not None:
        df["Month Year"] = df[month_candidate]
    else:
        df["Month Year"] = pd.Series(["" for _ in range(len(df))])

    for column in ["Employee Name", "Department", "Designation", "Unit_Raw", "Unit_Display", "State", "Month Year"]:
        if column in df.columns:
            df[column] = df[column].fillna("")

    df["Net Paid"] = pd.to_numeric(df.get("Net Paid", pd.Series([0 for _ in range(len(df))])), errors="coerce").fillna(0)

    for unit_column in ["Unit", "Address", "Address1", "Address_1"]:
        if unit_column in df_units.columns:
            df["Unit_Address"] = df_units[unit_column]
            break

    if "Unit_Clean" in df.columns and "Unit_Clean" in df_units.columns:
        # Use third column from Units_Master (contains unit names like "Ahemedabad", "Mumbai-Borivali")
        unit_name_col = df_units.columns[2] if len(df_units.columns) > 2 else "Unit"
        
        # Build lookup dictionary: Unit_Clean -> Unit Name
        if unit_name_col in df_units.columns and unit_name_col != "Unit_Clean":
            try:
                unit_lookup = df_units[["Unit_Clean", unit_name_col]].drop_duplicates(subset=["Unit_Clean"])
                unit_dict = dict(zip(unit_lookup["Unit_Clean"].astype(str).str.strip(), 
                                    unit_lookup[unit_name_col].astype(str).str.strip()))
                # Apply the lookup to Unit_Display
                df["Unit_Display"] = df["Unit_Clean"].map(unit_dict).fillna(df["Unit_Raw"].astype(str).str.strip())
            except Exception as e:
                st.warning(f"Unit name lookup issue: {e}")

    return df

def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return "".join(ch.lower() for ch in str(value) if ch.isalnum())


HEADER_KEYWORDS = ["employee", "code", "name", "father", "spouse", "designation", "department",
                    "days", "basic", "hra", "gross", "deduct", "net", "month", "period"]


def find_header_row(sheet) -> int:
    # A real header row has several SEPARATE short cells, each naming a
    # column (Code, Name, Father...). A title/preamble sentence lives in a
    # single merged cell and can accidentally contain one keyword substring
    # (e.g. "REGISTER OF WAGES OF EMPLOYEES" contains "employee"). Counting
    # distinct matching cells - not raw keyword occurrences - tells them
    # apart: require at least 2 separate header-bearing cells in the row.
    best_row = 10
    best_score = 0
    for row_idx in range(1, 26):
        row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 26)]
        normalized_values = [normalize_header(value) for value in row_values if value is not None]
        distinct_matching_cells = sum(
            1 for value in normalized_values if any(keyword in value for keyword in HEADER_KEYWORDS)
        )
        if distinct_matching_cells >= 2 and distinct_matching_cells > best_score:
            best_score = distinct_matching_cells
            best_row = row_idx
    return best_row


def get_unit_master_details(df_units: pd.DataFrame | None, selected_unit: str | None) -> dict[str, str]:
    if df_units is None or selected_unit is None:
        return {"unit_name": "", "unit_address": "", "ward_circle": ""}

    df_units = df_units.copy()
    df_units.columns = df_units.columns.astype(str).str.strip()

    if "Unit_Clean" not in df_units.columns:
        if "Unit" in df_units.columns:
            df_units["Unit_Clean"] = df_units["Unit"].apply(normalize_unit_key)
        else:
            return {"unit_name": "", "unit_address": "", "ward_circle": ""}

    clean_target = normalize_unit_key(selected_unit)
    match = df_units[df_units["Unit_Clean"].apply(normalize_unit_key) == clean_target]
    if match.empty:
        return {"unit_name": str(selected_unit).strip(), "unit_address": "", "ward_circle": ""}

    row = match.iloc[0]

    addr_col = next((col for col in df_units.columns if "address" in str(col).lower()), None)
    unit_address = ""
    if addr_col and pd.notna(row[addr_col]):
        unit_address = str(row[addr_col]).strip()
    elif len(row) > 4 and pd.notna(row.iloc[4]):
        unit_address = str(row.iloc[4]).strip()

    ward_col = next((col for col in df_units.columns if any(w in str(col).lower() for w in ["ward", "circle"])), None)
    ward_circle = ""
    if ward_col and pd.notna(row[ward_col]):
        ward_circle = str(row[ward_col]).strip()
    elif len(row) > 12 and pd.notna(row.iloc[12]):
        ward_circle = str(row.iloc[12]).strip()

    return {
        "unit_name": str(row["Unit"]).strip() if "Unit" in row.index else str(selected_unit).strip(),
        "unit_address": unit_address,
        "ward_circle": ward_circle,
    }


def get_unit_master_column_value(df_units: pd.DataFrame | None, selected_unit: str | None, column_letter: str) -> object:
    """Look up a Units_Master column by its real Excel letter (e.g. 'F' for
    Timing-From, 'M' for Ward/Circle) for the selected unit's row. Positional
    (letter) lookup is used instead of column names because Units_Master has
    duplicate header text ('From'/'To' appear twice, for Timing and Rest)."""
    if df_units is None or df_units.empty or not selected_unit or not column_letter:
        return ""
    df_units = df_units.copy()
    df_units.columns = df_units.columns.astype(str).str.strip()
    unit_col = next((c for c in df_units.columns if str(c).strip().lower().rstrip() == "unit"), None)
    if unit_col is None:
        return ""
    match = df_units[df_units[unit_col].apply(normalize_unit_key) == normalize_unit_key(selected_unit)]
    if match.empty:
        return ""
    col_idx0 = column_letter_to_index(column_letter) - 1
    if col_idx0 >= len(match.columns):
        return ""
    value = match.iloc[0, col_idx0]
    return "" if pd.isna(value) else value


def _lookup_col_ref_value(letter: str, df_col_ref: pd.DataFrame | None, row_data: dict) -> object:
    """Look up a Coloums letter (e.g. 'B', 'EN') and return that field's
    value from the employee row dict."""
    if not letter or df_col_ref is None or df_col_ref.empty:
        return None
    df_col_ref = df_col_ref.copy()
    df_col_ref.columns = df_col_ref.columns.astype(str).str.strip()
    if len(df_col_ref.columns) < 2:
        return None
    lookup = df_col_ref.iloc[:, 0].astype(str).str.strip().str.upper()
    values = df_col_ref.iloc[:, 1].astype(str).str.strip()
    match_idx = lookup[lookup == letter].index
    if len(match_idx) > 0:
        col_name = values.iloc[match_idx[0]]
        if col_name in row_data:
            return row_data.get(col_name, "")
    return None


COL_PREFIX_RE = re.compile(r"^col[\s\-_]+", re.IGNORECASE)


def _is_col_prefixed(rule_text: str) -> bool:
    return bool(COL_PREFIX_RE.match(rule_text.strip()))


def _extract_col_letter(token: str) -> str:
    """Pull the trailing Excel column letters out of a rule token - handles
    'Col_EN', 'Col-B', 'Col AO' (space-separated), and a bare 'EP' used on
    the right side of a '+' or with no prefix at all."""
    token = token.strip()
    token = COL_PREFIX_RE.sub("", token)
    return "".join(ch for ch in token if ch.isalpha()).upper()


def resolve_col_rule(
    rule_str: object,
    df_col_ref: pd.DataFrame | None,
    row_data: dict,
    idx: int = 0,
    active_month: str | None = None,
    selected_year: int | None = None,
    df_units: pd.DataFrame | None = None,
    selected_unit: str | None = None,
) -> object:
    if rule_str is None:
        return ""

    rule_text = str(rule_str).strip()
    if not rule_text:
        return ""

    lowered_rule = rule_text.lower()

    if "fixed text" in lowered_rule:
        # Handles every real variant seen in the master sheet: "(Fixed Text)",
        # "( Fixed Text)" (extra space), and "- Fixed Text" (no parens at all).
        cleaned = re.sub(r"[\(\-]?\s*fixed\s*text\s*\)?", "", rule_text, flags=re.IGNORECASE)
        return cleaned.strip(" -")

    if lowered_rule == "auto_generated":
        return idx + 1

    if lowered_rule == "wage month":
        month_value = active_month if active_month not in {"All", "", None} else "July"
        year_value = selected_year if selected_year is not None else datetime.now().year
        return f"{month_value[:3]}-{str(year_value)[-2:]}"

    active_month_name = active_month if active_month not in {"All", "", None} else "July"
    year_value = selected_year if selected_year is not None else datetime.now().year
    try:
        month_num = datetime.strptime(active_month_name[:3], "%b").month
    except ValueError:
        month_num = datetime.now().month
    days_in_month = calendar.monthrange(int(year_value), month_num)[1]

    if "1st day" in lowered_rule and "month" in lowered_rule:
        return datetime(int(year_value), month_num, 1).strftime("%d-%b-%Y")
    if "last day" in lowered_rule and "month" in lowered_rule:
        return datetime(int(year_value), month_num, days_in_month).strftime("%d-%b-%Y")
    if "days in cal" in lowered_rule:
        return days_in_month

    if "unit master" in lowered_rule and "col" in lowered_rule:
        # e.g. "Unit Master_Col_F" or the combined
        # "Unit Master_Col_C + Unit Master_Col_E" (Name + Address).
        parts = re.split(r"\s*\+\s*", rule_text)
        resolved_parts = []
        for part in parts:
            match = re.search(r"col_([a-z]+)", part.strip(), flags=re.IGNORECASE)
            if match:
                value = get_unit_master_column_value(df_units, selected_unit, match.group(1))
                if value not in (None, ""):
                    resolved_parts.append(str(value).strip())
        return ", ".join(resolved_parts)

    if _is_col_prefixed(rule_text) and "+" in rule_text:
        # Compound rule summing two source columns, e.g. "Col_EO+EP" (Earned
        # Leave = PL Accrued + PL Monthly Increment) or "Col-CB+CD"
        # (Payments Made = Salary Advance + Loan).
        total = 0.0
        any_value = False
        for part in rule_text.split("+"):
            letter = _extract_col_letter(part)
            value = _lookup_col_ref_value(letter, df_col_ref, row_data)
            if value not in (None, ""):
                try:
                    total += float(value)
                    any_value = True
                except (TypeError, ValueError):
                    pass
        return total if any_value else ""

    if _is_col_prefixed(rule_text):
        letter = _extract_col_letter(rule_text)
        value = _lookup_col_ref_value(letter, df_col_ref, row_data)
        return value if value is not None else ""

    if rule_text.isalpha() and rule_text.isupper() and 1 <= len(rule_text) <= 3:
        # A bare column-letter reference with no "Col" prefix at all, e.g.
        # Form V's day-by-day columns are just "CU", "CV", "CW"...
        value = _lookup_col_ref_value(rule_text, df_col_ref, row_data)
        if value is not None:
            return value

    return rule_text


def get_value_for_header(
    header_value: object,
    row: dict,
    unit_master_details: dict[str, str] | None = None,
    df_mapping_rules: pd.DataFrame | None = None,
    df_col_ref: pd.DataFrame | None = None,
    idx: int = 0,
    df_units: pd.DataFrame | None = None,
    selected_unit: str | None = None,
) -> object:
    normalized = normalize_header(header_value)

    if df_mapping_rules is not None and not df_mapping_rules.empty:
        for _, rule_row in df_mapping_rules.iterrows():
            values = [str(value).strip() for value in rule_row.dropna().tolist() if str(value).strip()]
            if len(values) < 2:
                continue
            if normalize_header(values[0]) == normalized:
                return resolve_col_rule(
                    values[1], df_col_ref, row, idx, df_units=df_units, selected_unit=selected_unit
                )

    if "code" in normalized:
        return row.get("Code", "")
    if any(keyword in normalized for keyword in ["srno", "slno", "sino", "serialno"]):
        return idx + 1
    if any(keyword in normalized for keyword in ["name", "employee"]) and "father" not in normalized and "spouse" not in normalized:
        return row.get("Employee Name", "")
    if "father" in normalized:
        return row.get("Father Name", "")
    if "spouse" in normalized:
        return row.get("Spouse Name", "")
    if "uan" in normalized:
        return row.get("UAN No", "")
    if "pf" in normalized and "no" in normalized:
        return row.get("PF No", "")
    if "esic" in normalized and "old" in normalized:
        return row.get("ESIC Old No", "")
    if "pan" in normalized:
        return row.get("PAN", "")
    if "join" in normalized or "appoint" in normalized:
        return row.get("Joining_Date", "")
    if "exit" in normalized:
        return row.get("Exit_Date", "")
    if "design" in normalized:
        return row.get("Designation", "")
    if "depart" in normalized:
        return row.get("Department", "")
    if "day" in normalized and "paid" in normalized:
        return row.get("Days Paid", "")
    if "day" in normalized and "present" in normalized:
        return row.get("Days Present", "")
    if "basic" in normalized:
        return row.get("Earned Basic", "")
    if "hra" in normalized:
        return row.get("Earned HRA", "")
    if "gross" in normalized:
        return row.get("Earned Gross Salary", "")
    if "prov" in normalized or "provident" in normalized:
        return row.get("Prov Fund", "")
    if "esic" in normalized:
        return row.get("ESIC", "")
    if "ptax" in normalized or "tax" in normalized and "professional" in normalized:
        return row.get("PTax", "")
    if "tds" in normalized:
        return row.get("TDS", "")
    if "deduct" in normalized or "totaldeduct" in normalized:
        return row.get("Total Deductions", "")
    if "net" in normalized or "salary" in normalized:
        return row.get("Net Paid", "")
    if "month" in normalized or "period" in normalized:
        return row.get("Month Year", "")
    return ""


def column_letter_to_index(column_letter: str) -> int:
    column_letter = column_letter.upper()
    total = 0
    for char in column_letter:
        total = total * 26 + (ord(char) - ord("A") + 1)
    return total


def safe_write(sheet, cell_coordinate: str, value) -> None:
    cell = sheet[cell_coordinate]
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell_coordinate in merged_range:
                top_left_coord = merged_range.coord.split(":")[0]
                sheet[top_left_coord].value = value
                return
    cell.value = value


TEMPLATE_MARKER_FILLS = {"FFBDD6EE", "FFFFFF00", "FF9CC2E5"}


def is_reference_cell(cell) -> bool:
    """True for cells filled with the template's yellow 'reference/constant'
    color - these hold values that are the same for every employee (shift
    timings, 'Monthly', 'Approved in HRMS', etc.) or are intentionally
    manual/exception-only fields, and must never be blanked out."""
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == "rgb":
        return fill.fgColor.rgb == "FFFFFF00"
    return False


def detect_table_end_row(sheet, start_row: int, max_scan: int = 500) -> int:
    """Find the real last pre-formatted employee row by scanning downward
    from start_row for the template's marker fill colors (blue 'to-fill'
    cells or yellow constants), instead of assuming a fixed row count.
    Stops after 5 consecutive rows carry no marker fill at all."""
    last_marked_row = start_row
    blank_streak = 0
    for row_idx in range(start_row, start_row + max_scan):
        if row_idx > sheet.max_row:
            break
        row_has_marker = False
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.type == "rgb" and fill.fgColor.rgb in TEMPLATE_MARKER_FILLS:
                row_has_marker = True
                break
        if row_has_marker:
            last_marked_row = row_idx
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= 5:
                break
    return last_marked_row


def copy_cell_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)


def write_header_month(sheet, selected_month: str, selected_year: int | None = None) -> None:
    target_keywords = ["month", "for the period ending", "wage month", "period"]
    if selected_month in {"All", "", None}:
        selected_month = "July"
    if selected_year is None:
        selected_year = datetime.now().year

    month_label = f"{selected_month}-{str(selected_year)[-2:]}"
    for row_idx in range(1, 6):
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            normalized_value = normalize_header(cell.value)
            if any(keyword in normalized_value for keyword in target_keywords):
                next_cell = sheet.cell(row=row_idx, column=col_idx + 1)
                if next_cell.value is None or normalize_header(str(next_cell.value)) != normalize_header(month_label):
                    safe_write(sheet, next_cell.coordinate, month_label)
                else:
                    safe_write(sheet, cell.coordinate, month_label)
                return


def inject_form_dates(sheet, form_name: str, selected_month: str, selected_year: int) -> None:
    if selected_month in {"All", "", None}:
        selected_month = "July"
    year_suffix = str(selected_year)[-2:]
    month_label = f"{selected_month[:3]}-{year_suffix}"

    if form_name == "Form A":
        # The date isn't in its own cell - it's a substring inside the A8
        # paragraph ("...shall take effect from (date) June-26"). Replace
        # just that trailing date token so the rest of the sentence survives.
        cell_a8 = sheet["A8"]
        current_text = str(cell_a8.value or "")
        marker = "shall take effect from (date)"
        marker_idx = current_text.find(marker)
        if marker_idx != -1:
            prefix = current_text[: marker_idx + len(marker)]
            new_text = f"{prefix} {selected_month}-{year_suffix}"
        else:
            new_text = current_text
        safe_write(sheet, "A8", new_text)
    elif form_name == "Form E":
        # Real cell is A5, a single combined string - not split D5/E5.
        safe_write(sheet, "A5", f"Wage Period- {selected_month}-{selected_year}")
    elif form_name == "Form IV":
        # Real cell is A4, not G5.
        safe_write(sheet, "A4", f"Month Ending -{selected_month.upper()} {year_suffix}")
    elif form_name == "Form V":
        safe_write(sheet, "A5", f"For the period ending - {selected_month}-{selected_year}")


def verify_generated_form(
    file_path: str | os.PathLike[str],
    form_name: str | None = None,
    known_codes: set[str] | None = None,
) -> list[str]:
    errors: list[str] = []
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
    except Exception as exc:
        return [f"Unable to open generated workbook for QA: {exc}"]

    form_key = str(form_name).strip().upper() if form_name else ""
    start_row = {"FORM A": 19, "FORM C": 9, "FORM D": 8, "FORM E": 9, "FORM IV": 10, "FORM V": 11}.get(form_key, 9)
    known_codes = known_codes or set()

    for row_idx in range(1, start_row):
        for col_idx in range(1, 15):
            value = str(sheet.cell(row=row_idx, column=col_idx).value or "")
            if not value.strip():
                continue
            if "ALL-" in value.upper() or "ALL 20" in value.upper():
                errors.append(f"Invalid date string '{value}' found at row {row_idx}, col {col_idx}")
            if "empl_code" in value.lower():
                errors.append(f"Leaked column-header text '{value}' found in header block at row {row_idx}, col {col_idx}")
            elif value.strip() in known_codes:
                errors.append(f"Leaked employee-code value '{value}' found in header block at row {row_idx}, col {col_idx}")

    name_value = str(sheet.cell(row=start_row, column=3).value or "").strip()
    if not name_value:
        errors.append(f"Employee Name missing at starting row {start_row}")

    for row_idx in range(start_row, min(start_row + 3, sheet.max_row + 1)):
        row_has_data = any(str(sheet.cell(row=row_idx, column=col_idx).value or "").strip() for col_idx in range(1, 8))
        if row_has_data:
            first_value = str(sheet.cell(row=row_idx, column=1).value or "").strip()
            second_value = str(sheet.cell(row=row_idx, column=2).value or "").strip()
            third_value = str(sheet.cell(row=row_idx, column=3).value or "").strip()
            if not any([first_value, second_value, third_value]):
                errors.append(f"Empty employee row detected at row {row_idx}")

    return errors


def find_form_rule_row(df_mapping_rules: pd.DataFrame | None, form_name: str | None) -> int | None:
    """Sheet1 stacks all 6 forms' rules with the form name in column B of
    each block (e.g. row 2 = Form A's rules, row 5 = Form C's, ...).
    Returns the DataFrame row index for the current form's block, or None."""
    if df_mapping_rules is None or df_mapping_rules.empty or form_name is None:
        return None
    if df_mapping_rules.shape[1] < 2:
        return None
    target = str(form_name).strip().upper()
    for row_idx in range(df_mapping_rules.shape[0]):
        candidate = df_mapping_rules.iloc[row_idx, 1]
        if pd.notna(candidate) and str(candidate).strip().upper() == target:
            return row_idx
    return None


def generate_dynamic_form(
    filtered_df: pd.DataFrame,
    template_source: io.BytesIO,
    selected_month: str,
    selected_year: int,
    form_name: str | None = None,
    df_mapping_rules: pd.DataFrame | None = None,
    df_col_ref: pd.DataFrame | None = None,
    df_units: pd.DataFrame | None = None,
    selected_unit: str | None = None,
) -> BytesIO:
    template_source.seek(0)
    workbook = load_workbook(template_source)
    sheet = workbook.active

    write_header_month(sheet, selected_month, selected_year)
    if form_name:
        inject_form_dates(sheet, form_name, selected_month, selected_year)

    header_row = find_header_row(sheet)
    header_cells = [sheet.cell(row=header_row, column=col_idx).value for col_idx in range(1, 51)]

    # Verified directly against the real template files: the first row below the
    # header/source-annotation block where blue "to-fill" cells begin.
    FORM_START_ROWS = {
        "FORM A": 19,
        "FORM C": 9,
        "FORM D": 8,
        "FORM E": 9,
        "FORM IV": 10,
        "FORM V": 11,
    }
    form_key = str(form_name).strip().upper() if form_name else ""
    start_row = FORM_START_ROWS.get(form_key, 9)
    table_end_row = detect_table_end_row(sheet, start_row)
    capacity = max(0, table_end_row - start_row + 1)
    row_count = min(len(filtered_df), capacity)
    if len(filtered_df) > capacity:
        st.warning(
            f"{form_name}: {len(filtered_df)} employees selected but the template only "
            f"has {capacity} pre-formatted rows. Only the first {capacity} were written - "
            f"extend the template's formatted rows to include everyone."
        )

    active_month = selected_month if selected_month not in {"All", "", None} else "July"
    month_label = f"{active_month[:3]}-{str(selected_year)[-2:]}"
    wage_month_columns = {"Form E": "E", "Form D": "F"}
    unit_master_details = get_unit_master_details(df_units, selected_unit)
    form_rule_row = find_form_rule_row(df_mapping_rules, form_name)

    data_col_count = min(sheet.max_column, 51)
    for offset in range(row_count):
        row = filtered_df.iloc[offset].to_dict()
        row_idx = start_row + offset
        for col_idx in range(1, data_col_count + 1):
            target_cell = sheet.cell(row=row_idx, column=col_idx)
            source_cell = sheet.cell(row=start_row, column=col_idx)
            copy_cell_style(source_cell, target_cell)

            # Yellow cells hold constants that are the same for every employee
            # (shift timings, "Monthly", "Approved in HRMS", "NIL") or are
            # exception-only manual fields - never overwrite them.
            if is_reference_cell(source_cell):
                continue

            # Some templates (Form IV's "Total earning" column) have a live
            # per-row formula already built in - never overwrite a formula
            # with a computed data value.
            if isinstance(target_cell.value, str) and target_cell.value.startswith("="):
                continue

            rule_value = ""
            if df_mapping_rules is not None and not df_mapping_rules.empty and form_rule_row is not None:
                rule_col_index = col_idx + 7
                if rule_col_index < df_mapping_rules.shape[1]:
                    candidate = df_mapping_rules.iloc[form_rule_row, rule_col_index]
                    if pd.notna(candidate) and str(candidate).strip():
                        rule_value = str(candidate).strip()

            if rule_value:
                # An explicit Sheet1 rule exists for this column - use it.
                cell_value = resolve_col_rule(
                    rule_value,
                    df_col_ref,
                    row,
                    offset,
                    active_month=active_month,
                    selected_year=selected_year,
                    df_units=df_units,
                    selected_unit=selected_unit,
                )
            else:
                # No Sheet1 rule: fall back to matching the column's header
                # text (e.g. "Code", "Father's/Husband's Name") against the
                # employee's actual data - NOT writing the header text itself.
                header_value = header_cells[col_idx - 1] if col_idx <= len(header_cells) else None
                cell_value = get_value_for_header(
                    header_value,
                    row,
                    unit_master_details=unit_master_details,
                    df_mapping_rules=df_mapping_rules,
                    df_col_ref=df_col_ref,
                    idx=offset,
                    df_units=df_units,
                    selected_unit=selected_unit,
                )
            safe_write(sheet, f"{get_column_letter(col_idx)}{row_idx}", cell_value)

        if form_name in wage_month_columns:
            month_col_letter = wage_month_columns[form_name]
            month_col_idx = column_letter_to_index(month_col_letter)
            month_cell = sheet.cell(row=row_idx, column=month_col_idx)
            copy_cell_style(sheet.cell(row=start_row, column=month_col_idx), month_cell)
            safe_write(sheet, month_cell.coordinate, month_label)

    for row_idx in range(start_row + row_count, table_end_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            safe_write(sheet, cell.coordinate, None)
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font()
            cell.alignment = Alignment()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


st.set_page_config(page_title="HR Compliance Engine", layout="wide")


def check_password() -> bool:
    if st.session_state.get("authenticated", False):
        return True

    st.title("🔒 HR Compliance Portal - Secure Login")
    st.caption("Please enter your organization access code to open the dashboard.")
    password_input = st.text_input("Access Code", type="password", key="login_code")
    if st.button("Log In", type="primary"):
        correct_code = st.secrets.get("app_security", {}).get("access_code", "HungerPangs2026")
        if password_input == correct_code:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("❌ Incorrect Access Code. Please contact HR IT.")

    return False


if not check_password():
    st.stop()

st.markdown(custom_css, unsafe_allow_html=True)

if "forms_generated" not in st.session_state:
    st.session_state.forms_generated = False

if "generated_forms" not in st.session_state:
    st.session_state.generated_forms = {}

header_html = """
<div style="display: flex; flex-direction: column; align-items: center; justify-content: center; margin-bottom: 1.5rem;">
    <img src="https://drive.google.com/thumbnail?id=1OpUw3MCFGLRs7GQ4xezk5ouqYTvy6yv9&sz=w1000"
         style="max-height: 120px; width: auto; object-fit: contain; margin-bottom: 12px; filter: drop-shadow(0 4px 6px rgba(0,0,0,0.1));">
    <h1 style="color: #1d4ed8; font-family: 'Inter', sans-serif; font-weight: 800; font-size: 2.1rem; margin: 0; padding: 0; text-align: center;">
        HR Statutory Compliance Engine
    </h1>
</div>
"""
st.markdown(header_html, unsafe_allow_html=True)

[df_conso, df_units, df_mapping_rules, df_col_ref, df_emp_master, df_leave_register, df_attendance_register] = load_master_data()
merged_df = build_merged_view()
filtered_df = merged_df.copy()

if "generated_zip" not in st.session_state:
    st.session_state.generated_zip = None

if "last_generated_files" not in st.session_state:
    st.session_state.last_generated_files = []


def reset_forms():
    st.session_state.forms_generated = False
    st.session_state.generated_forms = {}
    st.session_state.generated_zip = None
    st.session_state.last_generated_files = []


def save_form_to_archive(
    form_bytes: BytesIO,
    selected_state: str,
    selected_unit: str,
    selected_month: str,
    selected_year: int,
    form_name: str,
) -> tuple[Path, Path]:
    active_month = selected_month if selected_month != "All" else "July"
    active_month_year = f"{active_month}-{selected_year}"

    target_static_dir = STATIC_ARCHIVE_DIR / str(selected_state) / str(selected_unit) / active_month_year
    target_static_dir.mkdir(parents=True, exist_ok=True)
    static_file_path = target_static_dir / f"{selected_state}_{selected_unit}_{form_name}_{active_month_year}.xlsx"
    with open(static_file_path, "wb") as f:
        f.write(form_bytes.getvalue())

    if ONEDRIVE_ARCHIVE_DIR:
        target_od_dir = ONEDRIVE_ARCHIVE_DIR / str(selected_state) / str(selected_unit) / active_month_year
        target_od_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(static_file_path, target_od_dir / static_file_path.name)
        return static_file_path, target_od_dir

    return static_file_path, target_static_dir


def compile_statutory_forms(
    form_names: list[str],
    filtered_df: pd.DataFrame,
    selected_state: str,
    selected_unit: str,
    selected_month: str,
    selected_year: int,
) -> tuple[dict[str, bytes], list[str], Path | None]:
    generated: dict[str, bytes] = {}
    errors: list[str] = []
    state_urls = st.secrets.get("templates", {}).get(selected_state, {})
    archive_target_dir: Path | None = None
    if not state_urls:
        errors.append(f"No template URLs configured for {selected_state}")
        return generated, errors, archive_target_dir

    active_month = selected_month if selected_month not in {"All", "", None} else "July"

    for form_name in form_names:
        if form_name not in state_urls:
            errors.append(form_name)
            continue
        try:
            response = requests.get(state_urls[form_name], timeout=15)
            response.raise_for_status()
            template_bytes = io.BytesIO(response.content)
            form_bytes = generate_dynamic_form(
                filtered_df,
                template_bytes,
                active_month,
                int(selected_year),
                form_name=form_name,
                df_mapping_rules=df_mapping_rules,
                df_col_ref=df_col_ref,
                df_units=df_units,
                selected_unit=selected_unit,
            )

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                temp_file.write(form_bytes.getvalue())
                temp_path = temp_file.name

            known_codes = {
                str(c).strip()
                for c in filtered_df.get("Code", pd.Series(dtype=str)).tolist()
                if str(c).strip()
            }
            qa_errors = verify_generated_form(temp_path, form_name, known_codes=known_codes)
            os.remove(temp_path)
            if qa_errors:
                errors.append(f"{form_name} QA failed: {', '.join(qa_errors)}")
                continue

            file_path, target_dir = save_form_to_archive(
                form_bytes,
                selected_state,
                selected_unit,
                selected_month,
                int(selected_year),
                form_name,
            )
            archive_target_dir = target_dir
            generated[form_name] = form_bytes.getvalue()
        except Exception as exc:
            errors.append(f"{form_name} ({exc})")

    return generated, errors, archive_target_dir


def build_archive_zip(file_paths: list[Path]) -> BytesIO:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for path in file_paths:
            zip_file.write(path, arcname=path.name)
    zip_buffer.seek(0)
    return zip_buffer


def list_archive_records() -> pd.DataFrame:
    records = []
    for state_dir in sorted(STATIC_ARCHIVE_DIR.iterdir()):
        if not state_dir.is_dir():
            continue
        for unit_dir in sorted(state_dir.iterdir()):
            if not unit_dir.is_dir():
                continue
            for month_dir in sorted(unit_dir.iterdir()):
                if not month_dir.is_dir():
                    continue
                for archive_file in sorted(month_dir.glob("*.xlsx")):
                    relative_link = f"{STATIC_WEB_PREFIX}/{state_dir.name}/{unit_dir.name}/{month_dir.name}/{archive_file.name}"
                    records.append(
                        {
                            "Form Name": archive_file.stem.replace(f"{state_dir.name}_{unit_dir.name}_", ""),
                            "State / Region": state_dir.name,
                            "Outlet / Brand Unit": unit_dir.name,
                            "Wage Month": month_dir.name,
                            "OneDrive File Link": relative_link,
                        }
                    )
    return pd.DataFrame(records)


def clear_archive_directory(target_dir: Path) -> None:
    if target_dir.exists() and target_dir.is_dir():
        shutil.rmtree(target_dir)
    target_dir.mkdir(parents=True, exist_ok=True)


tab1, tab2 = st.tabs(["⚡ Statutory Form Generator", "📁 OneDrive Form Archive"])

with tab1:
    st.markdown("### Selection Filters")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        if "State" in merged_df.columns and not merged_df["State"].dropna().eq("").all():
            state_list = ["All"] + sorted(merged_df["State"].dropna().astype(str).unique().tolist())
            selected_state = st.selectbox("State / Region", state_list, key="selected_state", on_change=reset_forms)
        else:
            selected_state = "All"
            st.info("State / Region column not found in the workbook; showing all records.")

    with col2:
        if "Unit_Display" in merged_df.columns:
            if selected_state == "All":
                unit_options = sorted(merged_df["Unit_Display"].dropna().astype(str).unique().tolist())
            else:
                unit_options = sorted(
                    merged_df.loc[merged_df["State"].astype(str) == str(selected_state), "Unit_Display"]
                    .dropna()
                    .astype(str)
                    .unique()
                    .tolist()
                )
            selected_unit = st.selectbox("Outlet / Brand Unit", ["All"] + unit_options, key="selected_unit", on_change=reset_forms)
        else:
            selected_unit = "All"
            st.info("Unit column not found; using all records.")

    with col3:
        if "Month Year" in merged_df.columns:
            month_list = ["All"] + sorted(merged_df["Month Year"].dropna().astype(str).unique().tolist())
        else:
            month_list = ["All"]
        selected_month = st.selectbox("Month-Year", month_list, key="selected_month", on_change=reset_forms)

    with col4:
        current_year = datetime.now().year
        year_options = [str(year) for year in range(current_year - 2, current_year + 3)]
        selected_year = st.selectbox("Year", year_options, index=year_options.index(str(current_year)), key="selected_year", on_change=reset_forms)

    if selected_state != "All" and "State" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["State"].astype(str) == str(selected_state)]

    if selected_unit != "All" and "Unit_Display" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["Unit_Display"].astype(str) == str(selected_unit)]

    if selected_month not in {"All", "", None} and "Month Year" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["Month Year"].astype(str) == str(selected_month)]

    total_emps = len(filtered_df)
    kpi_html = f"""
    <div style="display: flex; gap: 20px; margin-bottom: 20px;">
        <div style="flex: 1; background: #ffffff; padding: 18px 24px; border-radius: 12px; border: 1px solid #e5e7eb; box-shadow: 0 4px 6px rgba(0,0,0,0.05);">
            <p style="margin: 0; color: #64748b; font-size: 0.85rem; font-weight: 600;">Total Employees</p>
            <h2 style="margin: 4px 0 0 0; color: #1d4ed8; font-size: 1.8rem; font-weight: 700;">{total_emps}</h2>
        </div>
    </div>
    """
    st.markdown(kpi_html, unsafe_allow_html=True)

    if selected_state == "All" or selected_unit == "All":
        st.warning("⚠️ Please select a specific State and Outlet/Brand Unit to enable form generation.")
    else:
        state_urls = st.secrets.get("templates", {}).get(selected_state, {})
        if state_urls:
            available_forms = sorted(state_urls.keys())
            st.info(f"Available statutory form templates for {selected_state}: {', '.join(available_forms)}")
        else:
            st.warning(f"No statutory form templates configured for {selected_state}.")

        gen_col1, gen_col2 = st.columns(2)
        with gen_col1:
            selected_form = st.selectbox(
                "Select Statutory Form",
                ["Form A", "Form C", "Form D", "Form E", "Form IV", "Form V"],
                key="selected_form",
            )
            if st.button("Generate Selected Form", key="generate_selected"):
                with st.status("⚡ Generating forms and pushing to cloud OneDrive...", expanded=True) as status:
                    status.write("📄 Compiling statutory form templates...")
                    generated, errors, archive_dir = compile_statutory_forms(
                        [selected_form],
                        filtered_df,
                        selected_state,
                        selected_unit,
                        selected_month,
                        int(selected_year),
                    )
                    status.write(f"☁️ Auto-pushed to OneDrive folder: `{archive_dir}`")
                    active_month = selected_month if selected_month not in {"All", "", None} else "July"
                    active_month_year = f"{active_month}-{selected_year}"
                    filename = f"{selected_state}_{selected_unit}_{selected_form}_{active_month_year}.xlsx"
                    if selected_form in generated:
                        status.write("📧 Dispatching file to Outlook for cloud OneDrive sync...")
                        email_success, email_msg = email_file_to_outlook(
                            generated[selected_form], filename
                        )
                        if email_success:
                            status.write(f"☁️ {email_msg}")
                        else:
                            status.caption(f"⚠️ {email_msg}")
                    status.update(label="✅ Form generated and synced to OneDrive successfully!", state="complete")
                active_month = selected_month if selected_month not in {"All", "", None} else "July"
                active_month_year = f"{active_month}-{selected_year}"
                if errors:
                    for error in errors:
                        st.warning(error)
                if generated:
                    st.success(f"✅ Saved to OneDrive: `{archive_dir}`")
                    form_content = generated[selected_form]
                    st.download_button(
                        label=f"📄 Download {selected_form} (.xlsx)",
                        data=form_content,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument-spreadsheetml.sheet",
                        key=f"dl_instant_{selected_form}",
                    )

        with gen_col2:
            if st.button("Generate All Forms (ZIP)", key="generate_bulk"):
                with st.status("⚡ Generating forms and pushing to cloud OneDrive...", expanded=True) as status:
                    status.write("📄 Compiling statutory form templates...")
                    all_forms = ["Form A", "Form C", "Form D", "Form E", "Form IV", "Form V"]
                    generated, errors, archive_dir = compile_statutory_forms(
                        all_forms,
                        filtered_df,
                        selected_state,
                        selected_unit,
                        selected_month,
                        int(selected_year),
                    )
                    status.write(f"☁️ Auto-pushed to OneDrive folder: `{archive_dir}`")
                    active_month = selected_month if selected_month not in {"All", "", None} else "July"
                    active_month_year = f"{active_month}-{selected_year}"
                    if generated:
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                            for form_name, content in generated.items():
                                zip_file.writestr(
                                    f"{selected_state}_{selected_unit}_{form_name}_{active_month_year}.xlsx",
                                    content,
                                )
                        zip_buffer.seek(0)
                        email_filename = f"HR_Compliance_{selected_state}_{selected_unit}_{active_month_year}.zip"
                        status.write("📧 Dispatching archive to Outlook for cloud OneDrive sync...")
                        email_success, email_msg = email_file_to_outlook(zip_buffer, email_filename)
                        if email_success:
                            status.write(f"☁️ {email_msg}")
                        else:
                            status.caption(f"⚠️ {email_msg}")
                    status.update(label="✅ All forms generated and synced to OneDrive successfully!", state="complete")
                active_month = selected_month if selected_month not in {"All", "", None} else "July"
                active_month_year = f"{active_month}-{selected_year}"
                if errors:
                    for error in errors:
                        st.warning(error)
                if generated:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for form_name, content in generated.items():
                            zip_file.writestr(
                                f"{selected_state}_{selected_unit}_{form_name}_{active_month_year}.xlsx",
                                content,
                            )
                    zip_buffer.seek(0)
                    st.session_state.forms_generated = True
                    st.session_state.generated_forms = generated
                    st.session_state.generated_zip = zip_buffer.getvalue()
                    st.success(f"✅ All forms generated and archived in OneDrive: `{archive_dir}`")
                    st.download_button(
                        label="📦 Download All Generated Forms (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name=f"HR_Compliance_{selected_state}_{selected_unit}_{active_month_year}.zip",
                        mime="application/zip",
                        key="dl_instant_zip",
                    )
                    with st.expander("Individual Form Downloads"):
                        for form_name, content in generated.items():
                            st.download_button(
                                label=f"Download {form_name}",
                                data=content,
                                file_name=f"{selected_state}_{selected_unit}_{form_name}_{active_month_year}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument-spreadsheetml.sheet",
                                key=f"dl_individual_{form_name}",
                            )
with tab2:
    st.markdown("### OneDrive Form Archive Browser")
    if st.button("🗑️ Clear All OneDrive Archived Forms", key="clear_all_archive_forms"):
        clear_archive_directory(STATIC_ARCHIVE_DIR)
        if ONEDRIVE_ARCHIVE_DIR:
            clear_archive_directory(ONEDRIVE_ARCHIVE_DIR)
        st.success("All OneDrive archived forms were removed. Refresh the app to confirm the archive is empty.")

    archived_states = sorted([d.name for d in STATIC_ARCHIVE_DIR.iterdir() if d.is_dir()])
    if not archived_states:
        st.info("No archived forms found in OneDrive.")
    else:
        selected_archive_state = st.selectbox("Select Archived State", archived_states, key="archive_state")
        archive_state_dir = STATIC_ARCHIVE_DIR / selected_archive_state
        archived_units = sorted([d.name for d in archive_state_dir.iterdir() if d.is_dir()])
        if not archived_units:
            st.info("No outlets found for the selected archived state.")
        else:
            selected_archive_unit = st.selectbox("Select Archived Outlet", archived_units, key="archive_unit")
            archive_unit_dir = archive_state_dir / selected_archive_unit
            archived_months = sorted([d.name for d in archive_unit_dir.iterdir() if d.is_dir()])
            if not archived_months:
                st.info("No wage-month folders found for the selected outlet.")
            else:
                selected_archive_month = st.selectbox("Select Month-Year", archived_months, key="archive_month")
                archive_month_dir = archive_unit_dir / selected_archive_month
                archived_files = sorted([f for f in archive_month_dir.glob("*.xlsx") if f.is_file()])
                if not archived_files:
                    st.info("No archived forms found for this selection.")
                else:
                    archive_df = pd.DataFrame(
                        [
                            {
                                "Form Name": archive_file.stem.replace(
                                    f"{selected_archive_state}_{selected_archive_unit}_", ""
                                ),
                                "State / Region": selected_archive_state,
                                "Outlet / Brand Unit": selected_archive_unit,
                                "Wage Month": selected_archive_month,
                                "OneDrive File Link": f"{STATIC_WEB_PREFIX}/{selected_archive_state}/{selected_archive_unit}/{selected_archive_month}/{archive_file.name}",
git rm --cached .streamlit/secrets.toml                            }
                            for archive_file in archived_files
                        ]
                    )

                    st.dataframe(
                        archive_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "OneDrive File Link": st.column_config.LinkColumn(
                                "OneDrive File Link",
                                display_text="📥 Open / Download Archived Form (.xlsx)",
                            )
                        },
                    )
